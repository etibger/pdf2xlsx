# -*- coding: utf-8 -*-
"""
Read the Order details xlsx and extract data from it to the  Order classes
"""
from openpyxl import load_workbook, Workbook
from collections import OrderedDict
from itertools import repeat

ORDER_START_TOKEN = 'Line Item:'
SIZE_BEGIN_TOKEN = 'Size'
SIZE_END_TOKEN = 'Total Qty:'


class StateError(Exception):
    def __init__(self, message):
        self.message = message


class GetOrderDetail:

    STATES = ["NO_ORDER", "STYLE_COLOR", "STYLE_NAME", "WAIT_SIZE", "SIZE", "FINISHED"]
    SCALE0 = ['5','5.5','6','6.5','7','7.5','8','8.5','9','9.5','10','10.5','11','11.5','12','12.5','13','13.5','14','14.5',]
    # SCALE0 = [str(int(k)) if k.is_integer() else str(k) for k in [x/10 for x in list(range(50, 150, 5))]]
    SCALE1 = ['10.5C','11C','11.5C','12C','12.5C','13C','13.5C','1Y','1.5Y','2Y', '2.5Y','3Y',]
    SCALE2 = ['2C','2.5C','3C','3.5C','4C','4.5C','5C','5.5C','6C','6.5C','7C','7.5C','8C','8.5C','9C','9.5C','10C',]
    # SCALE2 = [str(int(k))+'C' if k.is_integer() else str(k)+'C' for k in [x/10 for x in list(range(20, 105, 5))]]
    SCALE3 = ['3.5Y','4Y','4.5Y','5Y','5.5Y','6Y','6.5Y','7Y','7.5Y','8Y','8.5Y','9Y',]
    # SCALE3 = [str(int(k))+'Y' if k.is_integer() else str(k)+'Y' for k in [x/10 for x in list(range(35, 95, 5))]]
    SCALES = [SCALE0, SCALE1, SCALE2, SCALE3]

    def __init__(self, CRDate=None, ord_id=None, wholesale=None, ord_name=None, sizes=None, scale=None, gender=None):
        self.state = "NO_ORDER"
        self.CRDate = CRDate
        self.ord_id = ord_id
        self.wholesale = wholesale
        self.ord_name = ord_name
        self.sizes = sizes
        self.scale = scale
        self.gender = gender
        self.generated_scale = False

    def gen_scale(self, current_size):
        for s in self.SCALES:
            if current_size in s:
                if s is self.SCALE0:
                    if "women" in self.ord_name.lower():
                        self.gender = 'w'
                    else:
                        self.gender = 'm'
                self.sizes = OrderedDict(zip(s, repeat('')))
                return self.sizes

    def __call__(self, row):
        """
        gets a new row to parse
        """
        if self.state == "NO_ORDER":
            if row[0].value == ORDER_START_TOKEN:
                self.CRDate = row[4].value
                self.state = "STYLE_COLOR"
            return None

        if self.state == "STYLE_COLOR":
            self.ord_id = row[1].value
            self.wholesale = row[7].value
            self.state = "STYLE_NAME"
            return None

        if self.state == "STYLE_NAME":
            self.ord_name = row[1].value
            self.state = "WAIT_SIZE"
            return None

        if self.state == "WAIT_SIZE":
            if row[0].value == SIZE_BEGIN_TOKEN:
                self.state = "SIZE"
            return None

        if self.state == "SIZE":
            if row[0].value == SIZE_END_TOKEN:
                self.state = "FINISHED"
                return self
            if self.generated_scale is False:
                self.generated_scale = True
                self.gen_scale(row[0].value)
            self.sizes[row[0].value] = int(row[2].value)
            return None

        if self.state != "FINISHED":
            raise StateError("Undefined state found!: {state}".format(**self.__dict__))

    def __str__(self):
        return "{CRDate}, {ord_id}, {wholesale}, {ord_name}, {sizes}, {state}".format(**self.__dict__)

    def to_list(self):
        tmp = [self.ord_name, self.ord_id, self.wholesale, self.CRDate]
        tmp.extend([None, None, None, None])
        if self.gender != 'm':
            tmp.extend([None, None])
        tmp.extend(self.sizes.values())
        return tmp

    def header_to_list(self):
        tmp = [None, None, None, None, None, None, None]
        tmp.extend(self.sizes.keys())
        return tmp


def write_xlsx(orders, filename='../test/out8.xlsx'):
    wb = Workbook()
    ws = wb.active
    for order in orders:
        ws.append(order.to_list())
    wb.save(filename)


def read_xlsx( item_class, filename):
    wb = load_workbook(filename=filename, read_only=True)
    ws_names = wb.get_sheet_names()
    order_list = []
    for ws_name in ws_names:
        ws = wb[ws_name]
        i = j = 0
        item_instance = item_class()
        for row in ws.rows:
            tmp = item_instance(row)
            if tmp:
                order_list.append(tmp)
                item_instance = item_class()
    return order_list


def main():
    order_list = read_xlsx(GetOrderDetail,
                           '../test/lat8.xlsx')
    write_xlsx(order_list)
    print("Conversion Finished")

if __name__ == "__main__":
    main()
